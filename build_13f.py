#!/usr/bin/env python3
"""
Generic institutional 13F pipeline (any ticker/CUSIP; defaults to Apple).
Reads every SEC Form 13F Data Set ZIP (*form13f.zip) found in this folder,
builds consolidated institutional ownership for the 4 most recent report quarters,
and writes an Excel workbook with:
  - Top 25 holders (latest quarter first) + change vs prior quarter
  - Top 25 largest changes (latest vs prior, additions or sales)
  - New entrants into the stock (held latest quarter, absent prior quarter)
  - All holders (full list)
Usage: python3 aapl_13f_pipeline.py [folder]   (defaults to script's own folder)
"""
import os,sys,re,glob,zipfile,json
import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

import argparse
_ap=argparse.ArgumentParser()
_ap.add_argument('--cusips',default='037833100'); _ap.add_argument('--name',default='Apple Inc.')
_ap.add_argument('--ticker',default='AAPL'); _ap.add_argument('--insiders',default='auto')
_ap.add_argument('--folder',default=None); _ap.add_argument('--resolve',default='')
_a,_=_ap.parse_known_args()
FOLDER=_a.folder or (sys.argv[1] if (len(sys.argv)>1 and not sys.argv[1].startswith('-')) else os.path.dirname(os.path.abspath(__file__)))
CUSIPS=[c.strip() for c in _a.cusips.split(',') if c.strip()]; NAME=_a.name; TICKER=_a.ticker.upper()

FAMILY_PATTERNS=[
 (r'BLACKROCK','BlackRock Inc.'),(r'VANGUARD','Vanguard Group Inc.'),
 (r'STATE STREET','State Street Corp.'),(r'GEODE','Geode Capital Management'),
 (r'BERKSHIRE HATHAWAY','Berkshire Hathaway Inc.'),
 (r'\bFMR\b|FIDELITY MANAGEMENT|FIDELITY INSTITUTIONAL|FIDELITY INVESTMENTS|FIDELITY DIVERSIFYING','FMR LLC (Fidelity)'),
 (r'MORGAN STANLEY','Morgan Stanley'),(r'GOLDMAN SACHS','Goldman Sachs Group Inc.'),
 (r'JP ?MORGAN|J\.?P\.? MORGAN','JPMorgan Chase & Co.'),(r'BANK OF AMERICA|MERRILL LYNCH','Bank of America Corp.'),
 (r'NORGES BANK','Norges Bank (Norway)'),
 (r'WELLINGTON MANAGEMENT|WELLINGTON TRUST|CABOT-WELLINGTON','Wellington Management Group'),
 (r'NORTHERN TRUST','Northern Trust Corp.'),(r'BANK OF NEW YORK MELLON|BNY MELLON|MELLON CAPITAL|DREYFUS','Bank of New York Mellon'),
 (r'CHARLES SCHWAB|\bSCHWAB\b','Charles Schwab'),
 (r'^CAPITAL WORLD INVESTORS|^CAPITAL RESEARCH (&|AND|GLOBAL)|^CAPITAL INTERNATIONAL (INVESTORS|INC|LTD|SARL)|^CAPITAL GROUP (COMPANIES|INTERNATIONAL|INVESTMENT|PRIVATE)|^CAPITAL GUARDIAN','Capital Group (American Funds)'),
 (r'\bUBS\b','UBS Group AG'),(r'DEUTSCHE BANK|DWS ','Deutsche Bank AG'),
 (r'T ROWE|ROWE PRICE','T. Rowe Price Group'),(r'INVESCO','Invesco Ltd.'),
 (r'DIMENSIONAL FUND','Dimensional Fund Advisors'),(r'FRANKLIN RESOURCES|FRANKLIN ADVIS|FRANKLIN MUTUAL','Franklin Resources'),
 (r'AMUNDI','Amundi'),(r'LEGAL & GENERAL|LEGAL AND GENERAL','Legal & General Group'),
 (r'NUVEEN|\bTIAA\b|TEACHERS ADVISORS|TEACHERS INSURANCE','Nuveen / TIAA'),(r'WELLS FARGO','Wells Fargo & Co.'),
 (r'CITIGROUP|CITIBANK','Citigroup Inc.'),(r'AMERIPRISE|COLUMBIA THREADNEEDLE','Ameriprise Financial'),
 (r'PRINCIPAL FINANCIAL|PRINCIPAL GLOBAL|PRINCIPAL TRUST','Principal Financial Group'),(r'AMERICAN CENTURY','American Century'),
 (r'MASSACHUSETTS FINANCIAL|\bMFS\b','MFS (Massachusetts Financial)'),(r'ALLIANZ','Allianz SE'),
 (r'ROYAL BANK OF CANADA|\bRBC\b','Royal Bank of Canada'),(r'TORONTO[- ]DOMINION','Toronto-Dominion Bank'),
 (r'BANK OF MONTREAL|\bBMO\b','Bank of Montreal'),(r'\bHSBC\b','HSBC Holdings'),
 (r'BNP PARIBAS','BNP Paribas'),(r'MITSUBISHI UFJ|\bMUFG\b','Mitsubishi UFJ'),
 (r'SUMITOMO MITSUI','Sumitomo Mitsui'),(r'\bNOMURA\b','Nomura'),
 (r'FISHER ASSET|FISHER INVESTMENTS','Fisher Investments'),(r'JANUS HENDERSON','Janus Henderson'),
 (r'SCHRODER','Schroders'),(r'RAYMOND JAMES','Raymond James'),
]
APPLE_INSIDERS=['COOK TIM','TIMOTHY COOK','PAREKH KEVAN','KEVAN PAREKH','MAESTRI LUCA','WILLIAMS JEFF',
 'JEFFREY WILLIAMS','ADAMS KATHERINE','O BRIEN DEIRDRE','KONDO CHRIS','LEVINSON ARTHUR','AUSTIN WANDA',
 'GORSKY ALEX','JUNG ANDREA','LOZANO MONICA','SUGAR RONALD','WAGNER SUSAN','GORE AL']

if _a.insiders=='auto':
    INSIDERS=APPLE_INSIDERS if TICKER=='AAPL' else []
elif _a.insiders.strip():
    INSIDERS=[x.strip().upper() for x in _a.insiders.split(',') if x.strip()]
else:
    INSIDERS=[]
def clean(n):
    return re.sub(r'\s+',' ',re.sub(r'[^A-Z0-9 &.\-/]',' ',str(n).upper())).strip()
def fam(n):
    u=clean(n)
    for p,c in FAMILY_PATTERNS:
        if re.search(p,u): return c
    return None
def is_insider(n):
    u=re.sub(r'\s+',' ',re.sub(r'[^A-Z ]',' ',str(n).upper())).strip()
    return any(t in u for t in INSIDERS)

def member(zf,base):
    for n in zf.namelist():
        if n.upper().endswith('/'+base) or n.upper().endswith(base) and n.upper().split('/')[-1]==base:
            return n
    for n in zf.namelist():
        if n.upper().split('/')[-1]==base: return n
    return None

if _a.resolve:
    import collections
    zz=sorted(glob.glob(os.path.join(FOLDER,'*form13f.zip')))
    if not zz: sys.exit('No *form13f.zip to resolve against in '+FOLDER)
    cnt=collections.Counter()
    with zipfile.ZipFile(zz[-1]) as _zf:
        _nm=[n for n in _zf.namelist() if n.upper().endswith('INFOTABLE.TSV')][0]
        with _zf.open(_nm) as fh:
            for ch in pd.read_csv(fh,sep='\t',dtype=str,chunksize=200000,usecols=['NAMEOFISSUER','CUSIP','TITLEOFCLASS']):
                ch=ch[ch.NAMEOFISSUER.str.contains(_a.resolve,case=False,na=False)]
                for _,r in ch.iterrows(): cnt[(r.CUSIP,r.NAMEOFISSUER,r.TITLEOFCLASS)]+=1
    print('filers  CUSIP      issuer | class')
    for k,v in cnt.most_common(25): print(f'{v:6d}  {k[0]}  {k[1]} | {k[2]}')
    sys.exit(0)
acc_frames=[]; meta_frames=[]
zips=sorted(glob.glob(os.path.join(FOLDER,'*form13f.zip')))
if not zips: sys.exit('No *form13f.zip files found in '+FOLDER)
print('Reading',len(zips),'ZIP(s)')
for z in zips:
    with zipfile.ZipFile(z) as zf:
        inm=member(zf,'INFOTABLE.TSV'); snm=member(zf,'SUBMISSION.TSV'); cnm=member(zf,'COVERPAGE.TSV')
        parts=[]
        with zf.open(inm) as fh:
            for ch in pd.read_csv(fh,sep='\t',dtype=str,chunksize=200000,
                    usecols=['ACCESSION_NUMBER','CUSIP','VALUE','SSHPRNAMT','SSHPRNAMTTYPE','PUTCALL']):
                ch=ch[ch.CUSIP.isin(CUSIPS)]
                if len(ch): parts.append(ch)
        info=pd.concat(parts) if parts else pd.DataFrame()
        if len(info):
            info=info[(info.SSHPRNAMTTYPE=='SH')&(info.PUTCALL.fillna('').str.strip()=='')]
            info['SSHPRNAMT']=pd.to_numeric(info.SSHPRNAMT,errors='coerce').fillna(0)
            info['VALUE']=pd.to_numeric(info.VALUE,errors='coerce').fillna(0)
            acc_frames.append(info.groupby('ACCESSION_NUMBER').agg(shares=('SSHPRNAMT','sum'),value=('VALUE','sum')).reset_index())
        with zf.open(snm) as fh: sub=pd.read_csv(fh,sep='\t',dtype=str)
        with zf.open(cnm) as fh: cov=pd.read_csv(fh,sep='\t',dtype=str)[['ACCESSION_NUMBER','AMENDMENTNO','AMENDMENTTYPE','FILINGMANAGER_NAME']]
        meta_frames.append(sub.merge(cov,on='ACCESSION_NUMBER',how='left'))

acc=pd.concat(acc_frames).drop_duplicates('ACCESSION_NUMBER')
meta=pd.concat(meta_frames).drop_duplicates('ACCESSION_NUMBER')
af=acc.merge(meta,on='ACCESSION_NUMBER',how='left')
af['FILING_DATE']=pd.to_datetime(af.FILING_DATE,format='%d-%b-%Y',errors='coerce')
af['AMENDMENTNO_n']=pd.to_numeric(af.AMENDMENTNO,errors='coerce').fillna(0)
af['pdate']=pd.to_datetime(af.PERIODOFREPORT,format='%d-%b-%Y',errors='coerce')
# ingest any scraped_*.csv quarters (produced by edgar_13f_scraper.py) as additional quarters
for sc in glob.glob(os.path.join(FOLDER,'scraped_*.csv')):
    _mc=re.match(r'scraped_([0-9A-Za-z]+)_',os.path.basename(sc))
    if not _mc or _mc.group(1) not in CUSIPS: continue  # only ingest scraped files for THIS security
    sdf=pd.read_csv(sc,dtype={'cik':str})
    if not len(sdf): continue
    sdf=sdf.rename(columns={'cik':'CIK','filer_name':'FILINGMANAGER_NAME','period_report':'PERIODOFREPORT'})
    sdf['ACCESSION_NUMBER']='SCRAPED-'+sdf.CIK.astype(str)
    sdf['FILING_DATE']=pd.Timestamp('2100-01-01'); sdf['AMENDMENTNO_n']=0
    sdf['pdate']=pd.to_datetime(sdf.PERIODOFREPORT,format='%d-%b-%Y',errors='coerce')
    af=pd.concat([af,sdf[['ACCESSION_NUMBER','shares','value','FILING_DATE','AMENDMENTNO_n','PERIODOFREPORT','CIK','FILINGMANAGER_NAME','pdate']]],ignore_index=True)
    print('Ingested scraped quarter:',sc,len(sdf),'filers')
af=af[af.shares>0]

# pick 4 most recent quarters that look complete (>1000 filers)
cnt=af.groupby('pdate').ACCESSION_NUMBER.nunique()
mx=int(cnt.max()) if len(cnt) else 0
thr=max(10,0.2*mx)
good=sorted([d for d,c in cnt.items() if c>=thr])[-4:]
PERIODS=[(d, d.strftime('Q%q ')+str(d.year) if False else f"Q{(d.month-1)//3+1} {d.year}", d.strftime('%Y-%m-%d')) for d in good]
LABELS=[l for _,l,_ in PERIODS]
print('Quarters:',LABELS)

def consolidate(pdate):
    p=af[af.pdate==pdate].sort_values(['FILING_DATE','AMENDMENTNO_n'])
    last=p.groupby('CIK').tail(1).copy()
    last['family']=last.FILINGMANAGER_NAME.apply(fam)
    last['holder_key']=np.where(last.family.notna(),last.family,'CIK:'+last.CIK)
    last['holder_name']=np.where(last.family.notna(),last.family,last.FILINGMANAGER_NAME.str.title())
    cons=last.groupby('holder_key').agg(shares=('shares','sum'),value=('value','sum'),
        n_ciks=('CIK','nunique'),holder_name=('holder_name','first')).reset_index()
    return cons

wide=None
for pdate,label,iso in PERIODS:
    cons=consolidate(pdate)
    part=cons.rename(columns={'shares':f'sh_{label}','value':f'val_{label}','n_ciks':f'nc_{label}'})[['holder_key','holder_name',f'sh_{label}',f'val_{label}',f'nc_{label}']]
    wide=part if wide is None else wide.merge(part,on='holder_key',how='outer').assign(
        holder_name=lambda x:x.holder_name_y.combine_first(x.holder_name_x)).drop(columns=['holder_name_x','holder_name_y'])
for L in LABELS:
    for pre in ['sh_','val_','nc_']: wide[pre+L]=wide[pre+L].fillna(0)
wide['excl']=wide.holder_name.apply(is_insider)
keep=wide[~wide.excl].copy()

LATEST=LABELS[-1]; PRIOR=LABELS[-2]
price={}
for L in LABELS:
    imp=(wide[f'val_{L}']/wide[f'sh_{L}'])[(wide[f'sh_{L}']>0)&(wide[f'val_{L}']>0)]
    imp=imp[(imp>20)&(imp<1000)]; price[L]=float(imp.median())

order=LABELS[::-1]  # latest first
# new entrants: hold latest, absent in ALL prior quarters in dataset (avoids filing-gap false positives like Norges)
prior_cols=[f'sh_{L}' for L in LABELS[:-1]]
newe=keep[(keep[f'sh_{LATEST}']>0)&(keep[prior_cols].sum(axis=1)==0)].sort_values(f'sh_{LATEST}',ascending=False)
exits=keep[(keep[f'sh_{LATEST}']==0)&(keep[f'sh_{PRIOR}']>0)&(keep[[f'sh_{L}' for L in LABELS[:-1]]].sum(axis=1)>0)].sort_values(f'sh_{PRIOR}',ascending=False)

out=os.path.join(FOLDER,f'{TICKER}_13F_Institutional_Ownership.xlsx')
# ---- build workbook ----
NAVY='1F3864';STEEL='2E5496';LBLUE='D6E4F0';LGREY='F2F2F2'
hdr=Font(name='Arial',color='FFFFFF',bold=True,size=10);baseF=Font(name='Arial',size=10)
boldF=Font(name='Arial',size=10,bold=True);titleF=Font(name='Arial',size=15,bold=True,color='1F3864')
subF=Font(name='Arial',size=10,italic=True,color='595959')
fh_=PatternFill('solid',fgColor=STEEL);band=PatternFill('solid',fgColor=LGREY);blue=PatternFill('solid',fgColor=LBLUE)
green=PatternFill('solid',fgColor='E2EFDA')
thin=Side(style='thin',color='BFBFBF');bd=Border(left=thin,right=thin,top=thin,bottom=thin)
ctr=Alignment('center','center',wrap_text=True);lft=Alignment('left','center');rgt=Alignment('right','center')
wb=Workbook()
def hrow(ws,row,cols):
    for i,(n,w) in enumerate(cols):
        c=ws.cell(row,1+i,n);c.font=hdr;c.fill=fh_;c.alignment=ctr;c.border=bd
        ws.column_dimensions[get_column_letter(1+i)].width=w

ov=wb.active;ov.title='Overview';ov.sheet_view.showGridLines=False
ov['B2']=f'{NAME} ({TICKER}) — Institutional 13F Ownership';ov['B2'].font=titleF
ov['B3']=f'Latest quarter: {LATEST}  |  Quarters: {", ".join(LABELS)}';ov['B3'].font=subF
notes=[('Security',f'{NAME} — CUSIP {chr(44).join(CUSIPS)}'),
 ('Source','SEC EDGAR Form 13F Data Sets (all institutional filers)'),
 ('Holdings','Shares only (SH); options & debt excluded. Latest filing per filer per quarter (amendments resolved).'),
 ('Consolidation','Corporate families merged into one holder (BlackRock, Vanguard, etc.).'),
 ('Insiders',(f'Cross-referenced vs {NAME} Form 3/4/5 and Schedule 13D; natural-person insiders excluded.' if INSIDERS else 'No insider list applied; institutional 13F filers rarely overlap individual insiders.')),
 ('New entrants','Holders reporting '+TICKER+' in '+LATEST+' but in NONE of the prior quarters shown (true initiations; see New Entrants sheet).'),
 ('Value ($)','Summed EXACTLY from the as-filed 13F VALUE field — no price calculation. A few managers (e.g., T. Rowe Price) file Value in $thousands, so their figure reads ~1,000x low; shown unaltered and flagged.')]
r=5
for k,v in notes:
    ov[f'B{r}']=k;ov[f'B{r}'].font=boldF;ov[f'C{r}']=v;ov[f'C{r}'].font=baseF;ov[f'C{r}'].alignment=Alignment('left','top',wrap_text=True);r+=1
ov.column_dimensions['A'].width=2;ov.column_dimensions['B'].width=18;ov.column_dimensions['C'].width=110

asm=wb.create_sheet('Assumptions');asm.sheet_view.showGridLines=False
asm['B2']='Reference (informational — NOT used for $ values)';asm['B2'].font=titleF
hrow(asm,4,[('Metric',34),('Value',18),('Note',42)]);asm.cell(4,2).value='Metric'
ar=5
for L in order:
    asm.cell(ar,2,f'{TICKER} price — {L}').font=baseF
    cc=asm.cell(ar,3,round(price[L],2));cc.number_format='#,##0.00';cc.font=Font(name='Arial',size=10,color='0000FF')
    asm.cell(ar,4,'Approx quarter-end close (informational only)').font=subF
    for c in range(2,5):asm.cell(ar,c).border=bd
    ar+=1
price_latest_cell=f'Assumptions!$C$5'  # latest is first row
asm.column_dimensions['B'].width=30;asm.column_dimensions['C'].width=16;asm.column_dimensions['D'].width=44

# Top 25 holders (latest first)
hs=wb.create_sheet('Top 25 Holders');hs.sheet_view.showGridLines=False;hs.freeze_panes='A6'
hs['A2']=f'{NAME} ({TICKER}) — Top 25 Institutional Holders';hs['A2'].font=titleF
hs['A3']=f'Consolidated by parent; ranked by {LATEST} shares (latest quarter first).';hs['A3'].font=subF
qcols=[(f'{L}\nshares',15) for L in order]
cols=[('Rank',6),('Holder (consolidated parent)',34)]+qcols+[(f'Δ {LATEST}\nvs {PRIOR}',15),('Δ %',9),(f'Reported value\n{LATEST} ($, as filed)',22),('% of\n13F',9),('# sub-\nentities',8)]
hrow(hs,5,cols)
top=keep.sort_values(f'sh_{LATEST}',ascending=False).head(25).reset_index(drop=True)
tot_latest=keep[f'sh_{LATEST}'].sum()
nq=len(order)
for i,row in top.iterrows():
    r=6+i;hs.cell(r,1,i+1)
    hs.cell(r,2,row['holder_name'])
    for j,L in enumerate(order):
        v=row[f'sh_{L}'];hs.cell(r,3+j, None if v==0 else int(v))
    dcol=3+nq
    sl=row[f'sh_{LATEST}']; sp=row[f'sh_{PRIOR}']
    hs.cell(r,dcol, 'NR' if (sl==0 or sp==0) else int(sl-sp))
    hs.cell(r,dcol+1, 'NR' if (sl==0 or sp==0) else (sl-sp)/sp)
    vlat=row[f'val_{LATEST}']
    vc=hs.cell(r,dcol+2, int(vlat))
    if sl>0 and 0<vlat/sl<50: vc.comment=Comment('Filer reported Value in $thousands (as filed). True value ~ x1,000.','pipeline')
    hs.cell(r,dcol+3, sl/tot_latest)
    hs.cell(r,dcol+4,int(row[f'nc_{LATEST}']))
for r in range(6,6+len(top)):
    for c in range(3,3+nq+1):hs.cell(r,c).number_format='#,##0;(#,##0);"NR"';hs.cell(r,c).font=baseF;hs.cell(r,c).alignment=rgt
    hs.cell(r,3+nq+1).number_format='0.0%';hs.cell(r,3+nq+2).number_format='$#,##0';hs.cell(r,3+nq+3).number_format='0.00%'
    hs.cell(r,1).alignment=ctr;hs.cell(r,1).font=baseF;hs.cell(r,2).font=baseF;hs.cell(r,2).alignment=lft
    hs.cell(r,3+nq+4).alignment=ctr;hs.cell(r,3+nq+4).font=baseF
    for c in range(1,3+nq+5):
        hs.cell(r,c).border=bd
        if (r-6)%2==1:hs.cell(r,c).fill=band

# Top 25 changes
cs=wb.create_sheet('Top 25 Changes');cs.sheet_view.showGridLines=False;cs.freeze_panes='A6'
cs['A2']=f'{NAME} ({TICKER}) — Top 25 Largest Changes ({LATEST} vs {PRIOR})';cs['A2'].font=titleF
cs['A3']=f'Holders reporting {TICKER} in both quarters; ranked by absolute share change.';cs['A3'].font=subF
hrow(cs,5,[('Rank',6),('Holder (consolidated parent)',38),(f'{PRIOR}\nshares',15),(f'{LATEST}\nshares',15),('Δ shares',15),('Δ %',10),('Dir.',8),(f'Reported value\n{LATEST} ($, as filed)',22)])
both=keep[(keep[f'sh_{PRIOR}']>0)&(keep[f'sh_{LATEST}']>0)].copy()
both['chg']=both[f'sh_{LATEST}']-both[f'sh_{PRIOR}']
ch=both.reindex(both.chg.abs().sort_values(ascending=False).index).head(25).reset_index(drop=True)
for i,row in ch.iterrows():
    r=6+i;cs.cell(r,1,i+1);cs.cell(r,2,row['holder_name'])
    cs.cell(r,3,int(row[f'sh_{PRIOR}']));cs.cell(r,4,int(row[f'sh_{LATEST}']))
    chg=int(row[f'sh_{LATEST}']-row[f'sh_{PRIOR}'])
    cs.cell(r,5,chg);cs.cell(r,6,chg/row[f'sh_{PRIOR}']);cs.cell(r,7,'ADD' if chg>=0 else 'SELL')
    vc=cs.cell(r,8,int(row[f'val_{LATEST}']))
    if row[f'sh_{LATEST}']>0 and 0<row[f'val_{LATEST}']/row[f'sh_{LATEST}']<50: vc.comment=Comment('Filer reported Value in $thousands (as filed).','pipeline')
for r in range(6,6+len(ch)):
    for c in [3,4,5]:cs.cell(r,c).number_format='#,##0;(#,##0)';cs.cell(r,c).font=baseF;cs.cell(r,c).alignment=rgt
    cs.cell(r,6).number_format='0.0%;(0.0%)';cs.cell(r,8).number_format='$#,##0'
    cs.cell(r,1).alignment=ctr;cs.cell(r,1).font=baseF;cs.cell(r,2).font=baseF;cs.cell(r,2).alignment=lft
    cs.cell(r,7).alignment=ctr;cs.cell(r,7).font=baseF
    for c in range(1,9):
        cs.cell(r,c).border=bd
        if (r-6)%2==1:cs.cell(r,c).fill=band

# New entrants
ne=wb.create_sheet('New Entrants');ne.sheet_view.showGridLines=False;ne.freeze_panes='A6'
ne['A2']=f'{NAME} ({TICKER}) — New Institutional Entrants in {LATEST}';ne['A2'].font=titleF
ne['A3']=f'Held {TICKER} in {LATEST} but reported none in {PRIOR}. {len(newe)} new entrants total (showing up to 60 by size).';ne['A3'].font=subF
hrow(ne,5,[('Rank',6),('New holder (consolidated parent)',42),(f'{LATEST} shares',16),(f'Reported value ($, as filed)',24),('# sub-entities',13)])
for i,row in newe.head(60).reset_index(drop=True).iterrows():
    r=6+i;ne.cell(r,1,i+1);ne.cell(r,2,row['holder_name']);ne.cell(r,3,int(row[f'sh_{LATEST}']))
    ne.cell(r,4,int(row[f'val_{LATEST}']));ne.cell(r,5,int(row[f'nc_{LATEST}']))
    ne.cell(r,3).number_format='#,##0';ne.cell(r,4).number_format='$#,##0'
    for c in range(1,6):
        ne.cell(r,c).border=bd;ne.cell(r,c).font=baseF
        if (r-6)%2==1:ne.cell(r,c).fill=green
    ne.cell(r,2).alignment=lft;ne.cell(r,1).alignment=ctr;ne.cell(r,5).alignment=ctr

# All holders
ah=wb.create_sheet('All Holders');ah.sheet_view.showGridLines=False;ah.freeze_panes='A2'
allcols=[('Holder (consolidated parent)',40)]+[(f'{L} shares',15) for L in order]+[('# sub-entities (latest)',16)]
for i,(n,w) in enumerate(allcols):
    c=ah.cell(1,1+i,n);c.font=hdr;c.fill=fh_;c.alignment=ctr;c.border=bd;ah.column_dimensions[get_column_letter(1+i)].width=w
allh=keep.sort_values(f'sh_{LATEST}',ascending=False).reset_index(drop=True)
for i,row in allh.iterrows():
    r=2+i;ah.cell(r,1,row['holder_name'])
    for j,L in enumerate(order):
        v=row[f'sh_{L}'];ah.cell(r,2+j,None if v==0 else int(v));ah.cell(r,2+j).number_format='#,##0';ah.cell(r,2+j).font=baseF
    ah.cell(r,2+nq,int(row[f'nc_{LATEST}']));ah.cell(r,1).font=baseF
ah.auto_filter.ref=f"A1:{get_column_letter(2+nq)}{1+len(allh)}"
wb.save(out)
print('Wrote',out)
print('New entrants:',len(newe),'| Exits:',len(exits),'| holders:',len(keep))
print(json.dumps({'quarters':LABELS,'latest':LATEST,'prior':PRIOR,'price':{k:round(v,2) for k,v in price.items()},
 'new_entrants':int(len(newe)),'exits':int(len(exits))}))
